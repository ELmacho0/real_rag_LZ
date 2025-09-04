from __future__ import annotations
from typing import List, Tuple
from pathlib import Path
import json


def iter_excel_sheets(path: Path) -> List[Tuple[str, str]]:
    """Read an Excel file and yield each sheet as JSON string.

    Returns a list of tuples ``(sheet_name, sheet_json)`` where ``sheet_json``
    is a JSON serialized array-of-arrays representing the sheet values.  The
    import of ``openpyxl`` happens lazily so that environments without the
    dependency can still import this module without immediately failing.
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:  # pragma: no cover - optional dependency
        raise RuntimeError("openpyxl is required for Excel processing") from e

    wb = load_workbook(filename=path, data_only=True)
    sheets: List[Tuple[str, str]] = []
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        sheets.append((ws.title, json.dumps(rows, ensure_ascii=False)))
    return sheets
